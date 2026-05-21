from __future__ import annotations

from datetime import time
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TNR = 'Times New Roman'

FONT_TITLE  = Font(name=TNR, bold=True, size=14)
FONT_HEADER = Font(name=TNR, bold=True, size=12)
FONT_NORMAL = Font(name=TNR, size=12)
FONT_MERAH  = Font(name=TNR, size=12, color='FF0000')
FILL_HEADER  = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
FILL_WEEKEND = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L  = Alignment(horizontal='left',   vertical='center')
ALIGN_C0 = Alignment(horizontal='center', vertical='center')


def _header_cell(ws, coord, value):
    c = ws[coord]
    c.value = value
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.border = BORDER
    c.alignment = ALIGN_C
    return c


def format_waktu(t: time | None) -> str:
    return t.strftime('%H:%M:%S') if t else ''


# ---------------------------------------------------------------------------
# Sheet 1 — Data Mentah
# ---------------------------------------------------------------------------

def buat_sheet_data_mentah(wb: Workbook, data_mentah: pd.DataFrame) -> None:
    ws = wb.create_sheet(title='Data Mentah')

    for col_idx, col_name in enumerate(data_mentah.columns, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER
        c.alignment = ALIGN_C0

    for row_idx, row in enumerate(data_mentah.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx)
            try:
                c.value = '' if pd.isna(value) else value
            except (TypeError, ValueError):
                c.value = value
            c.font = FONT_NORMAL
            c.border = BORDER

    for col_idx in range(1, len(data_mentah.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


# ---------------------------------------------------------------------------
# Sheet 2 — Rekapitulasi
# ---------------------------------------------------------------------------

def buat_sheet_rekapitulasi(
    wb: Workbook,
    rekapitulasi: list[dict],
    bulan_tahun: str,
    li_refs: dict[int, dict],
) -> None:
    ws = wb.create_sheet(title='Rekapitulasi')

    # Title rows — row 1: company header, row 3: month (row 2 intentionally blank)
    for row, text in [
        (1, 'REKAPITULASI LAPORAN KEHADIRAN KARYAWAN TECTONA GROUP'),
        (3, f'BULAN {bulan_tahun.upper()}'),
    ]:
        ws.merge_cells(f'A{row}:M{row}')
        c = ws[f'A{row}']
        c.value = text
        c.font = FONT_TITLE
        c.alignment = ALIGN_C0

    # ---- Header rows 5–6 ----
    # Columns that merge rows 5+6 into a single label.
    # Border must be set on BOTH anchor (row 5) and slave (row 6) cells so the
    # full outline appears — openpyxl only reads the bottom border from row 6.
    span_two_rows = {
        'A': 'No',
        'B': 'Nama Karyawan',
        'C': 'Cuti',
        'D': 'Sakit',
        'E': 'Izin',
        'I': 'Jumlah\nDalam Menit',
        'K': 'Total',
        'M': 'Jumlah',
    }
    for kol, label in span_two_rows.items():
        ws[f'{kol}5'].value = label
        ws.merge_cells(f'{kol}5:{kol}6')
        c = ws[f'{kol}5']
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER
        c.alignment = ALIGN_C
        ws[f'{kol}6'].border = BORDER  # bottom border of merged area

    # F: "Dinas Lapangan" — single merged cell spanning rows 5–6 with wrap text
    ws['F5'].value = 'Dinas Lapangan'
    ws.merge_cells('F5:F6')
    c = ws['F5']
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.border = BORDER
    c.alignment = ALIGN_C          # ALIGN_C has wrap_text=True
    ws['F6'].border = BORDER       # bottom border of merged area

    # G5:H5 merged — "Terlambat Masuk"
    ws.merge_cells('G5:H5')
    _header_cell(ws, 'G5', 'Terlambat Masuk')
    ws['H5'].border = BORDER  # slave cell still needs border
    _header_cell(ws, 'G6', 'Jam')
    _header_cell(ws, 'H6', 'Menit')

    # J: two-row label (Denda/Menit / Rp.), not merged
    _header_cell(ws, 'J5', 'Denda/Menit')
    _header_cell(ws, 'J6', 'Rp.')

    # L: two-row label (Sangsi / 20 X), not merged
    _header_cell(ws, 'L5', 'Sangsi')
    _header_cell(ws, 'L6', '20 X')

    # ---- Data rows ----
    for idx, k in enumerate(rekapitulasi):
        r = 7 + idx
        pin = k['pin']
        refs = li_refs.get(pin, {})

        def li(field: str) -> str | None:
            addr = refs.get(field)
            return f"='Laporan Individual'!{addr}" if addr else None

        # No., Nama
        for col, val in [(1, idx + 1), (2, k['nama'])]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = FONT_NORMAL
            c.border = BORDER
            c.alignment = ALIGN_L if col == 2 else ALIGN_C0

        # C–F: Cuti, Sakit, Izin, Dinas — reference editable cells in Laporan Individual
        for col, field in [(3, 'cuti'), (4, 'sakit'), (5, 'izin'), (6, 'dinas')]:
            c = ws.cell(row=r, column=col, value=li(field))
            c.font = FONT_NORMAL
            c.border = BORDER
            c.alignment = ALIGN_C0

        # J: Denda/Menit — user fills manually
        c = ws.cell(row=r, column=10, value=None)
        c.font = FONT_NORMAL
        c.border = BORDER
        c.alignment = ALIGN_C0

        # I = total late minutes, pulled from Terlambat Masuk cell in Laporan Individual
        c = ws.cell(row=r, column=9)
        c.value = li('terlambat')
        c.font = FONT_NORMAL
        c.number_format = '#,##0'
        c.border = BORDER
        c.alignment = ALIGN_C0

        # G = full hours late  (INT(total_minutes / 60))
        c = ws.cell(row=r, column=7)
        c.value = f'=INT(I{r}/60)'
        c.font = FONT_NORMAL
        c.border = BORDER
        c.alignment = ALIGN_C0

        # H = remaining minutes  (MOD(total_minutes, 60))
        c = ws.cell(row=r, column=8)
        c.value = f'=MOD(I{r},60)'
        c.font = FONT_NORMAL
        c.border = BORDER
        c.alignment = ALIGN_C0

        # K = Total (formula)
        c = ws.cell(row=r, column=11)
        c.value = f'=IF(AND(I{r}<>"",J{r}<>""),I{r}*J{r},"-")'
        c.font = FONT_NORMAL
        c.number_format = '#,##0'
        c.border = BORDER
        c.alignment = ALIGN_C0

        # L = Sangsi multiplier
        c = ws.cell(row=r, column=12, value=20)
        c.font = FONT_NORMAL
        c.border = BORDER
        c.alignment = ALIGN_C0

        # M = Jumlah (formula)
        c = ws.cell(row=r, column=13)
        c.value = f'=IF(AND(K{r}<>"-",K{r}<>""),K{r}*L{r},"-")'
        c.font = FONT_NORMAL
        c.number_format = '#,##0'
        c.border = BORDER
        c.alignment = ALIGN_C0

    lebar = {
        'A': 5, 'B': 28, 'C': 8, 'D': 8, 'E': 8, 'F': 12,
        'G': 8, 'H': 8, 'I': 16, 'J': 14, 'K': 14, 'L': 10, 'M': 14,
    }
    for kol, w in lebar.items():
        ws.column_dimensions[kol].width = w


# ---------------------------------------------------------------------------
# Sheet 4 (hidden) — Data Harian
# ---------------------------------------------------------------------------

def buat_sheet_data_harian(
    wb: Workbook,
    laporan_individual: dict[int, dict],
    daftar_nama: list[str],
) -> str:
    SHEET_NAME = 'Data Harian'
    ws = wb.create_sheet(title=SHEET_NAME)

    # Column layout (A–N):
    # A=Nama, B=NoHari, C=Hari, D=Tanggal, E=JamKerja, F=JamMasuk, G=JamKeluar,
    # H=MenitTerlambat(number), I=JamTerlambat(string), J=CatatanOtomatis,
    # K=IsWeekend, L=PIN, M=Key, N=CatatanManual(user-editable)
    headers = [
        'Nama', 'No Hari', 'Hari', 'Tanggal', 'Jam Kerja',
        'Jam Masuk', 'Jam Keluar', 'Menit Terlambat', 'Jam Terlambat',
        'Catatan Otomatis', 'Is Weekend', 'PIN', 'Key', 'Catatan (Manual)',
    ]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER

    baris = 2
    for pin in sorted(laporan_individual.keys()):
        data = laporan_individual[pin]
        nama = data['nama']

        for idx_hari, d in enumerate(data['detail'], 1):
            ws.cell(row=baris, column=1,  value=nama)
            ws.cell(row=baris, column=2,  value=idx_hari)
            ws.cell(row=baris, column=3,  value=d['hari'])
            ws.cell(row=baris, column=4,  value=d['tanggal'].strftime('%d-%m-%Y'))
            ws.cell(row=baris, column=5,  value=d['jam_kerja'])
            ws.cell(row=baris, column=6,  value=format_waktu(d['jam_masuk']))
            ws.cell(row=baris, column=7,  value=format_waktu(d['jam_keluar']))
            ws.cell(row=baris, column=8,  value=d['menit_terlambat'])   # numeric for SUMIF
            ws.cell(row=baris, column=9,  value=d['jam_terlambat'])     # display string
            ws.cell(row=baris, column=10, value=d['catatan_otomatis'])
            ws.cell(row=baris, column=11, value=1 if d['is_weekend'] else 0)
            ws.cell(row=baris, column=12, value=data['pin'])
            ws.cell(row=baris, column=13, value=f"{nama}|{idx_hari}")  # lookup key
            ws.cell(row=baris, column=14, value='')                     # Catatan (Manual)

            for col in range(1, 15):
                ws.cell(row=baris, column=col).font = FONT_NORMAL
                ws.cell(row=baris, column=col).border = BORDER

            baris += 1

    # Auto-fit column widths based on max content length
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Auto-filter on header row
    ws.auto_filter.ref = f'A1:{get_column_letter(14)}{baris - 1}'

    return SHEET_NAME


# ---------------------------------------------------------------------------
# Sheet 3 — Laporan Individual (one stacked block per employee)
# ---------------------------------------------------------------------------

SIGNERS = [
    ('Dibuat Oleh :', 'A', 'C'),
    ('Diperiksa Oleh :', 'D', 'F'),
    ('Diketahui Oleh :', 'G', None),
]
SIGNER_NAMES = ['Justia Rifki Krismantara', 'Danta Putra Perdana', 'Hermanto Pribadi']

NOTES_LINES = [
    ('cuti',         ' Cuti                          ', 'Hari'),
    ('sakit',        ' Sakit                        ', 'Hari'),
    ('izin',         ' Izin                           ', 'Hari'),
    ('dinas',        ' Lapangan /Dinas     ', 'Hari'),
    ('tepat_waktu',  ' Tepat Waktu            ', 'Hari'),
    ('terlambat',    ' Terlambat Masuk    ', 'Menit'),
    ('cepat_pulang', ' Cepat Pulang', 'Menit'),
]

COL_HEADERS = ['Hari', 'Tanggal', 'Jam Kerja', 'Jam Masuk', 'Jam Keluar', 'Jam Terlambat', 'Catatan']


def buat_sheet_individual_static(
    wb: Workbook,
    laporan_individual: dict[int, dict],
) -> dict[int, dict]:
    """
    Writes one block per employee to 'Laporan Individual', stacked vertically.
    Returns {pin: {field: cell_address}} so Rekapitulasi can reference key cells.

    Block layout (R = block start row, n = days in month):
      R+0         title (merged A:G)
      R+1, R+2   blank
      R+3         Fingerprint ID
      R+4         Nama Karyawan
      R+5, R+6   column headers (each col A–G merged over 2 rows)
      R+7 … R+6+n  daily data rows
      R+7+n       TOTAL row
      R+8+n, R+9+n  blank
      R+10+n      signature label row
      R+11+n … R+13+n  blank (signature space)
      R+14+n      signer names
      R+15+n      blank
      R+16+n      "Catatan :"
      R+17+n      Cuti
      R+18+n      Sakit
      R+19+n      Izin
      R+20+n      Lapangan/Dinas
      R+21+n      Tepat Waktu (formula)
      R+22+n      Terlambat Masuk (formula)
      R+23+n      Cepat Pulang
      R+24+n … R+26+n  blank separator
      R+27+n      ← next block starts here
    """
    ws = wb.create_sheet(title='Laporan Individual')

    pins = sorted(laporan_individual.keys())
    li_refs: dict[int, dict] = {}

    R = 1
    for pin in pins:
        data = laporan_individual[pin]
        nama = data['nama']
        detail = data['detail']
        n = len(detail)

        ds = R + 7          # data start row
        de = R + 6 + n      # data end row
        r_total = R + 7 + n

        # ---- Title ----
        ws.merge_cells(f'A{R}:G{R}')
        c = ws[f'A{R}']
        c.value = 'LAPORAN KEHADIRAN KARYAWAN'
        c.font = FONT_TITLE
        c.alignment = ALIGN_C0

        # ---- ID / Name rows ----
        ws[f'A{R+3}'].value = 'Fingerprint ID'
        ws[f'A{R+3}'].font = FONT_HEADER
        ws[f'C{R+3}'].value = pin
        ws[f'C{R+3}'].font = FONT_NORMAL

        ws[f'A{R+4}'].value = 'Nama Karyawan'
        ws[f'A{R+4}'].font = FONT_HEADER
        ws[f'C{R+4}'].value = nama
        ws[f'C{R+4}'].font = FONT_NORMAL

        # ---- Column headers (double-row merged per column) ----
        r_hdr = R + 5
        for ci, label in enumerate(COL_HEADERS, 1):
            cl = get_column_letter(ci)
            ws.merge_cells(f'{cl}{r_hdr}:{cl}{r_hdr+1}')
            c = ws[f'{cl}{r_hdr}']
            c.value = label
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.border = BORDER
            c.alignment = ALIGN_C
            ws[f'{cl}{r_hdr+1}'].border = BORDER

        # ---- Daily data rows ----
        for idx, d in enumerate(detail):
            r = ds + idx
            is_wknd = d['is_weekend']
            font = FONT_MERAH if is_wknd else FONT_NORMAL
            fill = FILL_WEEKEND if is_wknd else None

            menit = d['menit_terlambat']
            row_vals = [
                ('A', d['hari'],                            ALIGN_C0),
                ('B', d['tanggal'],                         ALIGN_C0),
                ('C', d['jam_kerja'],                       ALIGN_C0),
                ('D', format_waktu(d['jam_masuk']),         ALIGN_C0),
                ('E', format_waktu(d['jam_keluar']),        ALIGN_C0),
                ('F', menit if menit > 0 else None,         ALIGN_C0),
                ('G', d['catatan_otomatis'],                ALIGN_L),
            ]
            for cl, val, align in row_vals:
                c = ws[f'{cl}{r}']
                c.value = val
                c.font = font
                c.border = BORDER
                c.alignment = align
                if fill:
                    c.fill = fill
            ws[f'B{r}'].number_format = 'DD-MM-YYYY'

        # ---- TOTAL row ----
        ws.merge_cells(f'A{r_total}:G{r_total}')
        c = ws[f'A{r_total}']
        c.value = 'TOTAL'
        c.font = FONT_HEADER
        c.alignment = ALIGN_C0
        c.border = BORDER

        # ---- Signature label row ----
        r_sig_lbl = r_total + 3
        ws.merge_cells(f'A{r_sig_lbl}:C{r_sig_lbl}')
        ws.merge_cells(f'D{r_sig_lbl}:F{r_sig_lbl}')
        for cell_ref, text in [
            (f'A{r_sig_lbl}', 'Dibuat Oleh :'),
            (f'D{r_sig_lbl}', 'Diperiksa Oleh :'),
            (f'G{r_sig_lbl}', 'Diketahui Oleh :'),
        ]:
            c = ws[cell_ref]
            c.value = text
            c.font = FONT_NORMAL
            c.alignment = ALIGN_L

        # ---- Signer names row ----
        r_sig_names = r_total + 7
        ws.merge_cells(f'A{r_sig_names}:C{r_sig_names}')
        ws.merge_cells(f'D{r_sig_names}:F{r_sig_names}')
        for cell_ref, text in [
            (f'A{r_sig_names}', 'Justia Rifki Krismantara'),
            (f'D{r_sig_names}', 'Danta Putra Perdana'),
            (f'G{r_sig_names}', 'Hermanto Pribadi'),
        ]:
            c = ws[cell_ref]
            c.value = text
            c.font = FONT_NORMAL
            c.alignment = ALIGN_L

        # ---- "Catatan :" header ----
        r_cat_hdr = r_total + 9
        ws[f'A{r_cat_hdr}'].value = 'Catatan :'
        ws[f'A{r_cat_hdr}'].font = FONT_HEADER

        # ---- Notes lines ----
        r_notes = r_total + 10
        li_refs[pin] = {}
        for i, (key, label, unit) in enumerate(NOTES_LINES):
            r_note = r_notes + i
            ws.merge_cells(f'A{r_note}:B{r_note}')
            c = ws[f'A{r_note}']
            c.value = label
            c.font = FONT_NORMAL
            c.alignment = ALIGN_L

            ws[f'C{r_note}'].value = ':'
            ws[f'C{r_note}'].font = FONT_NORMAL
            ws[f'C{r_note}'].alignment = ALIGN_C0

            c = ws[f'D{r_note}']
            if key == 'tepat_waktu':
                c.value = (
                    f'=COUNTIF(C{ds}:C{de},"08:00-17:00")'
                    f'-COUNTIF(F{ds}:F{de},">"&0)'
                )
            elif key == 'terlambat':
                c.value = f'=SUM(F{ds}:F{de})'
            c.font = FONT_NORMAL
            c.number_format = '0'
            c.alignment = ALIGN_C0

            ws[f'E{r_note}'].value = unit
            ws[f'E{r_note}'].font = FONT_NORMAL
            ws[f'E{r_note}'].alignment = ALIGN_L

            li_refs[pin][key] = f'D{r_note}'

        R += n + 27

    # ---- Column widths ----
    for kol, w in {'A': 22, 'B': 14, 'C': 14, 'D': 12, 'E': 12, 'F': 16, 'G': 30}.items():
        ws.column_dimensions[kol].width = w

    return li_refs


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def buat_file_excel(
    path_output: str,
    data_mentah: pd.DataFrame,
    rekapitulasi: list[dict],
    laporan_individual: dict[int, dict],
    bulan_tahun: str,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    daftar_nama = [
        laporan_individual[pin]['nama']
        for pin in sorted(laporan_individual.keys())
    ]

    buat_sheet_data_mentah(wb, data_mentah)
    li_refs = buat_sheet_individual_static(wb, laporan_individual)
    buat_sheet_rekapitulasi(wb, rekapitulasi, bulan_tahun, li_refs)
    buat_sheet_data_harian(wb, laporan_individual, daftar_nama)

    sheet_order = ['Data Mentah', 'Rekapitulasi', 'Laporan Individual', 'Data Harian']
    for i, name in enumerate(sheet_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.active = wb.sheetnames.index('Laporan Individual')
    wb.save(path_output)
